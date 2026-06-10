"""Ingestion orchestration (ORM port of the NestJS ingestion module).

Ports three NestJS providers that lived under ``apps/api/src/ingestion``:

  - ``DocumentTextService``  (document-text.service.ts) — extract plain text from
    docx / pdf / xlsx / pptx / plain-text buffers and resolve it into chunks,
    with a metadata-only fallback.
  - ``SyncCursorService``    (sync-cursor.service.ts)    — per user/drive/source
    delta-link cursors for OneDrive delta sync.
  - ``IngestionService``     (ingestion.service.ts)      — Outlook + document
    (SharePoint / OneDrive) sync into Pinecone, ingestion status, job creation
    and the worker sync-target listing.

CROSS-MODULE NOTE: the ``graph``, ``pinecone`` and ``rag`` sibling modules are
part of the rag/ingestion/query import cycle (and some are not migrated yet), so
their services are imported LAZILY inside the methods that use them. The ``ai``
and ``audit`` services are not part of the cycle and are imported normally.
"""

import json
import os
import re
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app import logger
from app.core.constants import EMBEDDING_DIMENSIONS
from app.core.settings import settings
from app.core.utils import chunk_text, vector_id
from app.models.audit import AuditLog
from app.models.ingestion import IngestionJob, SyncCursor
from app.schema.auth import AuthContext
from app.schema.common import (
    DocumentSourceSyncResult,
    DocumentsSyncResult,
    IngestionStatus,
    IngestionVerification,
    OutlookSyncResult,
)
from app.schema.ingestion import ExtractResult, WorkerSyncTarget
from app.services.ai_client import AiClientService
from app.services.audit_service import AuditService

_PLAIN_TEXT_EXTENSIONS = frozenset(
    {".txt", ".md", ".csv", ".json", ".xml", ".html", ".htm", ".log"}
)

_SKIP_EXTENSIONS = frozenset(
    {
        ".png",
        ".jpg",
        ".jpeg",
        ".gif",
        ".bmp",
        ".svg",
        ".ico",
        ".webp",
        ".mp4",
        ".mp3",
        ".wav",
        ".zip",
        ".rar",
        ".7z",
        ".exe",
        ".dll",
        ".bin",
    }
)

_PPTX_TEXT_RE = re.compile(r"<a:t[^>]*>([^<]*)</a:t>")
_WS_RE = re.compile(r"[ \t]+")


def _extension(file_name: str) -> str:
    """Return the lowercased file extension including the dot (or '')."""
    i = file_name.rfind(".")
    if i < 0:
        return ""
    return file_name[i:].lower()


class DocumentTextService:
    """Extract plain text from file buffers and resolve into chunks.

    Port of the NestJS ``DocumentTextService``.
    """

    def resolve_chunks(
        self,
        file_name: str,
        web_url: str | None,
        text_content: str | None,
        skip_reason: str | None = None,
    ) -> ExtractResult:
        if skip_reason:
            return self._metadata_fallback(file_name, web_url, skip_reason)

        trimmed = (text_content or "").strip()
        if trimmed:
            full = chunk_text(f"{file_name}\n\n{trimmed}")
            if full:
                return ExtractResult(chunks=full, reason="ok")

        return self._metadata_fallback(file_name, web_url, "empty_extracted_text")

    def extract_from_buffer(
        self,
        file_name: str,
        buffer: bytes,
        mime_type: str | None = None,
    ) -> dict[str, Any]:
        """Return ``{"text": str | None, "skip_reason": str | None}``."""
        ext = _extension(file_name)

        if ext in _SKIP_EXTENSIONS:
            return {"text": None, "skip_reason": f"unsupported_extension:{ext}"}

        if ext in _PLAIN_TEXT_EXTENSIONS or (mime_type and mime_type.startswith("text/")):
            return {"text": buffer.decode("utf-8", errors="replace"), "skip_reason": None}

        if (
            ext == ".docx"
            or mime_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            try:
                return {"text": _extract_docx_text(buffer), "skip_reason": None}
            except Exception as err:
                logger.warning(
                    f"[document-text] docx failed file={file_name}: {err}"
                )
                return {"text": None, "skip_reason": "docx_parse_failed"}

        if ext == ".pdf" or mime_type == "application/pdf":
            try:
                extracted = _extract_pdf_text(buffer)
                if not extracted.strip():
                    # Parsed fine but no text layer — almost always a scanned PDF.
                    logger.warning(
                        f"[document-text] pdf has no extractable text layer "
                        f"file={file_name} (scanned image?)"
                    )
                    return {"text": None, "skip_reason": "pdf_no_text_layer"}
                return {"text": extracted, "skip_reason": None}
            except Exception as err:
                logger.warning(f"[document-text] pdf failed file={file_name}: {err}")
                return {"text": None, "skip_reason": "pdf_parse_failed"}

        if (
            ext in (".xlsx", ".xls")
            or mime_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            try:
                return {"text": _extract_xlsx_text(buffer) or None, "skip_reason": None}
            except Exception as err:
                logger.warning(f"[document-text] xlsx failed file={file_name}: {err}")
                return {"text": None, "skip_reason": "xlsx_parse_failed"}

        if (
            ext == ".pptx"
            or mime_type
            == "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        ):
            try:
                return {"text": _extract_pptx_text(buffer) or None, "skip_reason": None}
            except Exception as err:
                logger.warning(f"[document-text] pptx failed file={file_name}: {err}")
                return {"text": None, "skip_reason": "pptx_parse_failed"}

        return {
            "text": None,
            "skip_reason": f"unsupported_type:{ext or mime_type or 'unknown'}",
        }

    def _metadata_fallback(
        self, file_name: str, web_url: str | None, reason: str
    ) -> ExtractResult:
        lines = [
            f"[Document: {file_name}]",
            f"URL: {web_url}" if web_url else "",
            "Indexed metadata only — file content could not be extracted.",
        ]
        chunks = chunk_text("\n".join(line for line in lines if line))
        if chunks:
            logger.info(
                f"[document-text] metadata fallback file={file_name} reason={reason}"
            )
            return ExtractResult(chunks=chunks, reason=f"metadata_fallback:{reason}")

        return ExtractResult(chunks=[], reason=f"no_chunks:{reason}")


def _extract_docx_text(buffer: bytes) -> str:
    """Extract raw text from a .docx buffer via python-docx."""
    from docx import Document  # lazy: optional parsing dependency

    document = Document(BytesIO(buffer))
    return "\n".join(p.text for p in document.paragraphs)


def _extract_pdf_text(buffer: bytes) -> str:
    """Extract text from a PDF buffer via pypdf."""
    from pypdf import PdfReader  # lazy: optional parsing dependency

    reader = PdfReader(BytesIO(buffer))
    parts: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        page_text = _WS_RE.sub(" ", page_text).strip()
        if page_text:
            parts.append(page_text)
    return "\n\n".join(parts)


def _extract_xlsx_text(buffer: bytes) -> str:
    """Extract CSV-ish text from an .xlsx buffer via openpyxl."""
    from openpyxl import load_workbook  # lazy: optional parsing dependency

    workbook = load_workbook(BytesIO(buffer), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell) for cell in row]
            rows.append(",".join(cells))
        csv = "\n".join(rows)
        if csv.strip():
            parts.append(f"Sheet: {sheet.title}\n{csv}")
    workbook.close()
    return "\n\n".join(parts)


def _extract_pptx_text(buffer: bytes) -> str:
    """Extract slide text from a .pptx buffer (matches the NestJS XML scan)."""
    parts: list[str] = []
    with zipfile.ZipFile(BytesIO(buffer)) as archive:
        slide_names = sorted(
            name
            for name in archive.namelist()
            if re.match(r"ppt/slides/slide\d+\.xml$", name, re.IGNORECASE)
        )
        for name in slide_names:
            xml = archive.read(name).decode("utf-8", errors="replace")
            texts = _PPTX_TEXT_RE.findall(xml)
            if texts:
                parts.append(" ".join(texts))
    return "\n\n".join(parts)


class SyncCursorService:
    """Per user/drive/source delta-link cursors (port of SyncCursorService)."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def get_delta_link(
        self, ctx: AuthContext, source: str, drive_id: str
    ) -> str | None:
        result = await self.db.execute(
            select(SyncCursor.delta_link).where(
                SyncCursor.organization_id == ctx.organization_id,
                SyncCursor.user_id == ctx.user_id,
                SyncCursor.source == source,
                SyncCursor.drive_id == drive_id,
            )
        )
        return result.scalar_one_or_none()

    async def set_delta_link(
        self,
        ctx: AuthContext,
        source: str,
        drive_id: str,
        delta_link: str | None,
    ) -> None:
        stmt = pg_insert(SyncCursor).values(
            organization_id=ctx.organization_id,
            user_id=ctx.user_id,
            source=source,
            drive_id=drive_id,
            delta_link=delta_link,
        )
        stmt = stmt.on_conflict_do_update(
            index_elements=[
                SyncCursor.organization_id,
                SyncCursor.user_id,
                SyncCursor.source,
                SyncCursor.drive_id,
            ],
            set_={
                "delta_link": stmt.excluded.delta_link,
                "updated_at": text("NOW()"),
            },
        )
        await self.db.execute(stmt)
        await self.db.commit()


class IngestionService:
    """Outlook + document ingestion into Pinecone (port of IngestionService)."""

    def __init__(self, db: AsyncSession, cache: Any | None = None) -> None:
        self.db = db
        self.cache = cache
        self.ai = AiClientService()
        self.audit = AuditService(db)
        self.document_text = DocumentTextService()
        self.sync_cursors = SyncCursorService(db)

    # --- lazy sibling-service accessors (graph / pinecone / rag cycle) ----

    def _graph(self) -> Any:
        from app.services.graph_service import GraphService

        return GraphService(self.db)

    def _pinecone(self) -> Any:
        from app.services.pinecone_service import PineconeService

        return PineconeService()

    def _outlook_mail(self) -> Any:
        from app.services.outlook_mail_service import OutlookMailService

        return OutlookMailService(self.db)

    def _sharepoint_docs(self) -> Any:
        from app.services.sharepoint_documents_service import SharePointDocumentsService

        return SharePointDocumentsService(self.db)

    # --- worker sync targets ----------------------------------------------

    async def list_worker_sync_targets(self) -> list[WorkerSyncTarget]:
        """Users with Microsoft tokens, for the scheduled worker sync."""
        rows = (
            await self.db.execute(
                text(
                    """
                    SELECT om.user_id, om.organization_id, om.role::text AS role
                    FROM organization_members om
                    INNER JOIN oauth_tokens ot ON ot.user_id = om.user_id
                    ORDER BY om.organization_id, om.user_id
                    """
                )
            )
        ).all()
        return [
            WorkerSyncTarget(
                user_id=str(r.user_id),
                organization_id=str(r.organization_id),
                role=r.role,
            )
            for r in rows
        ]

    # --- ingestion jobs ----------------------------------------------------

    async def create_job(
        self, ctx: AuthContext, job_type: str, payload: dict[str, Any]
    ) -> IngestionJob:
        job = IngestionJob(
            organization_id=ctx.organization_id,
            user_id=ctx.user_id,
            job_type=job_type,
            status="pending",
            payload=payload,
        )
        self.db.add(job)
        await self.db.commit()
        await self.db.refresh(job)
        return job

    # --- status ------------------------------------------------------------

    async def get_status(self, ctx: AuthContext) -> IngestionStatus:
        counts = (
            await self.db.execute(
                text(
                    """
                    SELECT
                      COUNT(*) AS email_metadata_count,
                      COUNT(*) FILTER (WHERE is_indexed = true) AS indexed_email_count,
                      (SELECT COUNT(*) FROM file_metadata
                       WHERE organization_id = :org AND user_id = :usr)
                        AS file_metadata_count,
                      (SELECT COUNT(*) FILTER (WHERE is_indexed = true) FROM file_metadata
                       WHERE organization_id = :org AND user_id = :usr)
                        AS indexed_file_count
                    FROM email_metadata
                    WHERE organization_id = :org AND user_id = :usr
                    """
                ),
                {"org": ctx.organization_id, "usr": ctx.user_id},
            )
        ).first()

        last_outlook_sync = await self._last_sync_at(ctx, "ingestion.outlook.sync")
        last_documents_sync = await self._last_sync_at(ctx, "ingestion.documents.sync")
        last_outlook_error = await self._last_sync_error(ctx, "ingestion.outlook.sync")
        last_documents_error = await self._last_sync_error(ctx, "ingestion.documents.sync")

        metadata_only_file_count = 0
        try:
            row = (
                await self.db.execute(
                    text(
                        """
                        SELECT COUNT(*) AS count FROM file_metadata
                        WHERE organization_id = :org AND user_id = :usr
                          AND is_indexed = true
                          AND index_reason LIKE 'metadata_fallback%'
                        """
                    ),
                    {"org": ctx.organization_id, "usr": ctx.user_id},
                )
            ).first()
            metadata_only_file_count = int(row.count) if row else 0
        except Exception:
            metadata_only_file_count = 0

        pinecone = self._pinecone()
        pinecone_vector_count = await pinecone.get_namespace_vector_count(
            ctx.organization_id
        )
        index_name = pinecone.get_index_name()

        scheduled_sync_interval_ms = settings.ingestion_sync_interval_ms
        scheduled_sync_enabled = bool(
            settings.worker_service_token and settings.worker_service_token.strip()
        )
        last_sync_ms = max(
            _parse_ms(last_outlook_sync),
            _parse_ms(last_documents_sync),
        )
        estimated_next_scheduled_sync_at: str | None = None
        if scheduled_sync_enabled and last_sync_ms > 0:
            estimated_next_scheduled_sync_at = (
                datetime.fromtimestamp(
                    (last_sync_ms + scheduled_sync_interval_ms) / 1000.0,
                    tz=timezone.utc,
                )
                .isoformat()
                .replace("+00:00", "Z")
            )

        return IngestionStatus(
            organization_id=ctx.organization_id,
            pinecone_index_name=index_name,
            pinecone_namespace=ctx.organization_id,
            pinecone_vector_count=pinecone_vector_count,
            email_metadata_count=int(counts.email_metadata_count) if counts else 0,
            indexed_email_count=int(counts.indexed_email_count) if counts else 0,
            file_metadata_count=int(counts.file_metadata_count) if counts else 0,
            indexed_file_count=int(counts.indexed_file_count) if counts else 0,
            last_outlook_sync_at=last_outlook_sync,
            last_documents_sync_at=last_documents_sync,
            last_outlook_sync_error=last_outlook_error,
            last_documents_sync_error=last_documents_error,
            metadata_only_file_count=metadata_only_file_count,
            ai_service_reachable=await self.ai.is_reachable(),
            scheduled_sync_enabled=scheduled_sync_enabled,
            scheduled_sync_interval_ms=scheduled_sync_interval_ms,
            estimated_next_scheduled_sync_at=estimated_next_scheduled_sync_at,
            verification=IngestionVerification(
                check_namespace_in_pinecone_console=(
                    f'Open index "{index_name}" → Namespaces → select '
                    f'"{ctx.organization_id}" (not __default__)'
                ),
                expected_dimensions=EMBEDDING_DIMENSIONS,
            ),
        )

    async def _last_sync_at(self, ctx: AuthContext, action: str) -> str | None:
        row = (
            await self.db.execute(
                select(AuditLog.created_at)
                .where(
                    AuditLog.organization_id == ctx.organization_id,
                    AuditLog.user_id == ctx.user_id,
                    AuditLog.action == action,
                )
                .order_by(AuditLog.created_at.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        return _iso(row)

    async def _last_sync_error(self, ctx: AuthContext, action: str) -> str | None:
        row = (
            await self.db.execute(
                select(AuditLog.metadata_)
                .where(
                    AuditLog.organization_id == ctx.organization_id,
                    AuditLog.user_id == ctx.user_id,
                    AuditLog.action == action,
                    AuditLog.metadata_["success"].astext == "false",
                )
                .order_by(AuditLog.created_at.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        if isinstance(row, dict):
            return row.get("error")
        return None

    # --- outlook sync ------------------------------------------------------

    async def sync_outlook(self, ctx: AuthContext) -> OutlookSyncResult:
        try:
            return await self._sync_outlook_inner(ctx)
        except Exception as err:
            message = str(err)
            await self.audit.log_audit(
                ctx,
                "ingestion.outlook.sync",
                "ingestion",
                None,
                {"success": False, "error": message},
            )
            raise

    async def _sync_outlook_inner(self, ctx: AuthContext) -> OutlookSyncResult:
        graph = self._graph()
        pinecone = self._pinecone()
        outlook_mail = self._outlook_mail()

        emails = await graph.fetch_recent_emails(ctx, 100)
        indexed = 0
        skipped = 0
        vectors_upserted = 0

        logger.info(
            f"[ingestion] outlook sync start org={ctx.organization_id} "
            f"emails={len(emails)}"
        )

        for email in emails:
            meta_id = await outlook_mail.upsert_email_metadata(ctx, email)
            content_id = meta_id or email["id"]

            chunks, reason = self._resolve_email_chunks(email)
            if not chunks:
                skipped += 1
                logger.warning(
                    f'[ingestion] skip email graphId={email["id"]} '
                    f'subject="{email.get("subject")}" reason={reason}'
                )
                continue

            stable_ids = [vector_id("outlook", content_id, str(i)) for i in range(len(chunks))]
            await pinecone.delete_by_vector_ids(ctx.organization_id, stable_ids)

            embed = await self.ai.embed(chunks)
            embeddings = embed["embeddings"]
            vectors = [
                pinecone.build_upsert(
                    "outlook",
                    content_id,
                    str(i),
                    embeddings[i],
                    {
                        "source": "outlook",
                        "organizationId": ctx.organization_id,
                        "emailId": content_id,
                        "sender": email.get("sender"),
                        "timestamp": email.get("receivedAt"),
                        "subject": email.get("subject"),
                        "text": chunk,
                    },
                )
                for i, chunk in enumerate(chunks)
            ]

            vectors_upserted += await pinecone.upsert(ctx.organization_id, vectors)

            if meta_id:
                await self.db.execute(
                    text(
                        "UPDATE email_metadata "
                        "SET is_indexed = true, indexed_at = NOW() WHERE id = :id"
                    ),
                    {"id": meta_id},
                )
                await self.db.commit()

            indexed += 1

        result = OutlookSyncResult(
            indexed=indexed,
            skipped=skipped,
            vectors_upserted=vectors_upserted,
            emails_fetched=len(emails),
            namespace=ctx.organization_id,
            pinecone_index_name=pinecone.get_index_name(),
        )

        await self.audit.log_audit(
            ctx,
            "ingestion.outlook.sync",
            "ingestion",
            None,
            {**result.model_dump(by_alias=True), "success": True},
        )

        logger.info(
            f"[ingestion] outlook sync done {json.dumps(result.model_dump(by_alias=True))}"
        )
        return result

    def _resolve_email_chunks(self, email: dict[str, Any]) -> tuple[list[str], str]:
        subject = email.get("subject") or ""
        body = email.get("body") or ""
        full = chunk_text(f"{subject}\n\n{body}")
        if full:
            return full, "ok"

        subject_only = chunk_text(subject.strip() or "(no subject)")
        if subject_only:
            logger.info(
                f'[ingestion] subject-only fallback graphId={email.get("id")} '
                f'subject="{subject}"'
            )
            return subject_only, "subject_fallback"

        return [], "empty_subject_and_body"

    # --- document sync -----------------------------------------------------

    async def sync_all_documents(self, ctx: AuthContext) -> DocumentsSyncResult:
        schema_ok = await self._has_file_metadata_drive_columns()
        if not schema_ok:
            from fastapi import HTTPException, status

            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database schema outdated for document sync. Run: pnpm db:migrate",
            )

        try:
            sharepoint = await self._sync_document_source(ctx, "sharepoint")
            onedrive = await self._sync_document_source(ctx, "onedrive")

            await self.audit.log_audit(
                ctx,
                "ingestion.documents.sync",
                "ingestion",
                None,
                {
                    "sharepoint": sharepoint.model_dump(by_alias=True),
                    "onedrive": onedrive.model_dump(by_alias=True),
                    "success": True,
                },
            )
            return DocumentsSyncResult(sharepoint=sharepoint, onedrive=onedrive)
        except Exception as err:
            from fastapi import HTTPException

            if isinstance(err, HTTPException):
                raise
            message = str(err)
            await self.audit.log_audit(
                ctx,
                "ingestion.documents.sync",
                "ingestion",
                None,
                {"success": False, "error": message},
            )
            raise

    async def sync_documents(
        self, ctx: AuthContext, source: str
    ) -> DocumentSourceSyncResult:
        """Deprecated: use ``sync_all_documents``."""
        return await self._sync_document_source(ctx, source)

    async def _sync_document_source(
        self, ctx: AuthContext, source: str
    ) -> DocumentSourceSyncResult:
        graph = self._graph()
        pinecone = self._pinecone()

        use_delta = source == "onedrive" and os.environ.get("ENABLE_DELTA_SYNC") != "false"

        if source == "sharepoint":
            files = await graph.collect_sharepoint_files(ctx)
        elif use_delta:
            files = await graph.collect_onedrive_files_delta(ctx, self.sync_cursors)
        else:
            files = await graph.collect_onedrive_files(ctx)

        indexed = 0
        skipped = 0
        vectors_upserted = 0
        limits = graph.get_crawl_limits()
        max_file_bytes = limits["maxFileBytes"] if isinstance(limits, dict) else limits.max_file_bytes

        logger.info(
            f"[ingestion] {source} sync start org={ctx.organization_id} "
            f"files={len(files)}"
        )

        for file in files:
            try:
                result = await self._index_drive_file(ctx, file, max_file_bytes)
                if result["indexed"]:
                    indexed += 1
                    vectors_upserted += result["vectors_upserted"]
                else:
                    skipped += 1
                    logger.warning(
                        f'[ingestion] skip file graphId={file.get("id")} '
                        f'name="{file.get("name")}" reason={result["reason"]}'
                    )
            except Exception as err:
                skipped += 1
                logger.warning(
                    f'[ingestion] skip file graphId={file.get("id")} '
                    f'name="{file.get("name")}" error={err}'
                )

        audit_action = (
            "ingestion.sharepoint.sync"
            if source == "sharepoint"
            else "ingestion.onedrive.sync"
        )

        sync_result = DocumentSourceSyncResult(
            indexed=indexed,
            skipped=skipped,
            vectors_upserted=vectors_upserted,
            files_fetched=len(files),
            namespace=ctx.organization_id,
            pinecone_index_name=pinecone.get_index_name(),
        )

        await self.audit.log_audit(
            ctx,
            audit_action,
            "ingestion",
            None,
            {**sync_result.model_dump(by_alias=True)},
        )

        logger.info(
            f"[ingestion] {source} sync done "
            f"{json.dumps(sync_result.model_dump(by_alias=True))}"
        )
        return sync_result

    async def _index_drive_file(
        self, ctx: AuthContext, file: dict[str, Any], max_file_bytes: int
    ) -> dict[str, Any]:
        graph = self._graph()
        pinecone = self._pinecone()
        sharepoint_docs = self._sharepoint_docs()

        meta_id = await sharepoint_docs.upsert_file_metadata(ctx, file)
        content_id = meta_id or file["id"]

        size = file.get("size")
        if size is not None and size > max_file_bytes:
            extract_result = self.document_text.resolve_chunks(
                file["name"], file.get("webUrl"), None, "file_too_large"
            )
        else:
            try:
                download = await graph.download_drive_item(
                    ctx, file.get("driveId"), file["id"]
                )
                buffer = download["buffer"] if isinstance(download, dict) else download.buffer
                extracted = self.document_text.extract_from_buffer(
                    file["name"], buffer, file.get("mimeType")
                )
                extract_result = self.document_text.resolve_chunks(
                    file["name"],
                    file.get("webUrl"),
                    extracted["text"],
                    extracted["skip_reason"],
                )
            except Exception as err:
                extract_result = self.document_text.resolve_chunks(
                    file["name"],
                    file.get("webUrl"),
                    None,
                    f"download_failed:{err}",
                )

        if not extract_result.chunks:
            return {
                "indexed": False,
                "vectors_upserted": 0,
                "reason": extract_result.reason,
            }

        stable_ids = [
            vector_id(file["source"], content_id, str(i))
            for i in range(len(extract_result.chunks))
        ]
        await pinecone.delete_by_vector_ids(ctx.organization_id, stable_ids)

        embed = await self.ai.embed(extract_result.chunks)
        embeddings = embed["embeddings"]
        vectors = [
            pinecone.build_upsert(
                file["source"],
                content_id,
                str(i),
                embeddings[i],
                {
                    "source": file["source"],
                    "organizationId": ctx.organization_id,
                    "fileId": content_id,
                    "fileName": file["name"],
                    "timestamp": file.get("lastModified"),
                    "webUrl": file.get("webUrl"),
                    "text": chunk,
                },
            )
            for i, chunk in enumerate(extract_result.chunks)
        ]

        upserted = await pinecone.upsert(ctx.organization_id, vectors)

        if meta_id:
            await self.db.execute(
                text(
                    "UPDATE file_metadata "
                    "SET is_indexed = true, indexed_at = NOW(), index_reason = :reason "
                    "WHERE id = :id"
                ),
                {"id": meta_id, "reason": extract_result.reason},
            )
            await self.db.commit()

        return {
            "indexed": True,
            "vectors_upserted": upserted,
            "reason": extract_result.reason,
        }

    async def _has_file_metadata_drive_columns(self) -> bool:
        row = (
            await self.db.execute(
                text(
                    """
                    SELECT COUNT(*) = 3 AS ok FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'file_metadata'
                      AND column_name IN ('drive_id', 'site_id', 'mime_type')
                    """
                )
            )
        ).first()
        return bool(row and row.ok)


def _iso(value: datetime | None) -> str | None:
    """Serialize a datetime to an ISO-8601 string (or None)."""
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.isoformat().replace("+00:00", "Z")


def _parse_ms(iso_value: str | None) -> int:
    """Parse an ISO-8601 string into epoch milliseconds (0 when absent/invalid)."""
    if not iso_value:
        return 0
    try:
        raw = iso_value.replace("Z", "+00:00")
        return int(datetime.fromisoformat(raw).timestamp() * 1000)
    except ValueError:
        return 0
